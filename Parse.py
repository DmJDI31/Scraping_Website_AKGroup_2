import json
import re
import os
import openpyxl
from openpyxl import load_workbook

import requests
from bs4 import BeautifulSoup

headers = {
    "Accept": "*/*",
    "User-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 YaBrowser/23.7.5.734 Yowser/2.5 Safari/537.36"
}

main_url = "https://www.retail.ru/rbc/tradingnetworks/"

req = requests.get(main_url, headers=headers)

src = req.text

soup = BeautifulSoup(src, "lxml")

all_segments_with_urls = soup.find(class_="row default-list").find_all("a")

all_segments_dict = {}
all_names_of_segments = []
for item in all_segments_with_urls:
    item_segment_name_text = item.text.strip()
    item_segment_urls_text = "https://www.retail.ru" + item.get("href")
    all_segments_dict[item_segment_name_text] = item_segment_urls_text
    all_names_of_segments.append(item_segment_name_text)

# запись в json-файл

with open("all_segments_with_urls/all_segments_with_urls.json", "w", encoding='utf-8') as file:
    json.dump(all_segments_dict, file, indent=4, ensure_ascii=False)

# проверка данных json-файла

with open("all_segments_with_urls/all_segments_with_urls.json", encoding='utf-8') as file:
    all_segment = json.load(file)

book = openpyxl.Workbook()
s = ""
table_head = ['Название компании',
              'Телефон',
              'E-mail',
              'Сайт',
              'VK',
              'YouTube',
              'О компании',
              'Общая информация']
hrefs = []
sheets_name = []
for segment_name, segment_href in all_segment.items():
    req = requests.get(url=segment_href, headers=headers)
    src = req.text
    soup = BeautifulSoup(src, "lxml")
    try:
        href = soup.find("li", class_="active").find("a")
        items_hrefs = "https://www.retail.ru" + href.get("href")
        hrefs.append(items_hrefs + "&PAGEN_1=")
    except Exception:
        items_hrefs = segment_href
        hrefs.append(items_hrefs)
    if len(segment_name) > 31:
        s = segment_name[:-(len(segment_name) - 31)]
    else:
        s = segment_name

    sheets_name.append(s)
    workbook = book.create_sheet(f"{s}")
del book['Sheet']

book.save("РОССИЯ.xlsx")

wb = load_workbook("РОССИЯ.xlsx")

for sheet in wb.worksheets:
    sheet['A1'] = table_head[0]
    sheet['B1'] = table_head[1]
    sheet['C1'] = table_head[2]
    sheet['D1'] = table_head[3]
    sheet['E1'] = table_head[4]
    sheet['F1'] = table_head[5]
    sheet['G1'] = table_head[6]
    sheet['H1'] = table_head[7]

wb.save("РОССИЯ.xlsx")

number_of_urls = 1
iteration_for = 1
count_of_pagination = []

for items_hrefs in hrefs:

    urls_pages_need = str(items_hrefs) + str(number_of_urls)
    req = requests.get(urls_pages_need, headers=headers)

    # req возвращает результат работы метода get библиотеки requests, а именно принимает первым аргументом url,
    # вторым (но он не обязателен) аргументом заголовки

    src = req.text

    # сохраняем в переменную наш полученный объект и вызовем метод .text

    # сохраняем полученные данные в файл

    with open(
            f"pages_of_segment/{sheets_name[iteration_for - 1]}_{number_of_urls}.html",
            "w", encoding="utf-8") as file:
        file.write(src)

    with open(
            f"pages_of_segment/{sheets_name[iteration_for - 1]}_{number_of_urls}.html",
            encoding="utf-8") as file:
        src = file.read()

    # передаём в объект 'soup'

    soup = BeautifulSoup(src, "lxml")
    try:
        href_of_pagination = soup.find("div", class_="pagination-wrap").find_all("a")[-1].get("href")
        print(href_of_pagination)
        if href_of_pagination[-2:] == '11':
            count_of_pagination.append(int(href_of_pagination[-2:]))
        count_of_pagination.append(int(href_of_pagination[-1:]))
    except Exception:
        count_of_pagination.append(1)

    os.remove(f"pages_of_segment/{sheets_name[iteration_for - 1]}_{number_of_urls}.html")
    iteration_for += 1

count_of_pagination.pop(7)
print(count_of_pagination)

for items_hrefs, item_sheets_name, cn_pagination in zip(hrefs, sheets_name, count_of_pagination):

    row = 2

    for number_of_urls_in_this_segment in range(1, cn_pagination + 1):

        urls_pages = str(items_hrefs) + str(number_of_urls_in_this_segment)

        req = requests.get(urls_pages, headers=headers)
        src = req.text

        with open(
                f"pages_of_segment/{item_sheets_name}_{number_of_urls_in_this_segment}.html",
                "w", encoding="utf-8") as file:
            file.write(src)

        with open(
                f"pages_of_segment/{item_sheets_name}_{number_of_urls_in_this_segment}.html",
                encoding="utf-8") as file:
            src = file.read()

        soup = BeautifulSoup(src, "lxml")
        all_companies = soup.find_all(class_="title")
        all_companies_hrefs = soup.find_all(class_="details")

        # zip для одновременного затрагивания несколько переменных
        all_companies_dict = {}
        for item_companies, item_href_companies in zip(all_companies, all_companies_hrefs):
            item_companies_text = item_companies.text.strip("\n")
            item_companies_href_text = "https://www.retail.ru" + item_href_companies.get("href")
            all_companies_dict[item_companies_text] = item_companies_href_text

        with open(
                f"data/all_href_companies_of_pages_{item_sheets_name}_{number_of_urls_in_this_segment}.json",
                "w",
                encoding='utf-8') as file:
            json.dump(all_companies_dict, file, indent=4, ensure_ascii=False)

        with open(
                f"data/all_href_companies_of_pages_{item_sheets_name}_{number_of_urls_in_this_segment}.json",
                encoding='utf-8') as file:
            all_companies = json.load(file)

        # узнаём количество страниц (на одной из pages) в данной категории
        # указываем int, т.к. возвращает объект строки

        iteration_count = int(len(all_companies))
        print(f'Всего итераций {iteration_count}')
        count = 0

        for company_name, company_href in all_companies.items():

            rep = ["\"", "*"]
            for item in rep:
                if item in company_name:
                    company_name = company_name.replace(item, '')
            req = requests.get(url=company_href, headers=headers)
            src = req.text
            # запись страниц в папку data. Счётчик count для перебора этих страниц
            with open(f"data/{company_name}.html", "w", encoding="utf-8") as file:
                file.write(src)

            with open(f"data/{company_name}.html", encoding="utf-8") as file:
                src = file.read()

            soup = BeautifulSoup(src, "lxml")

            # собираем данные о компаниях (почта, номер и т.д.)

            try:
                companies_name = soup.find(class_="col-lg-9 col-md-8 left-colom").find('h1')
            except Exception:
                companies_name = ""
            if None is soup.find(class_="noShowPhone showPhone"):
                col_number = ""
            else:
                col_number = "Показать телефон"
            e_mail = soup.select("a.prop_item.email")
            if e_mail is None:
                e_mail = ""
            else:
                e_mail = BeautifulSoup("".join(map(str, soup.select("a.prop_item.email"))), "html.parser")
            try:
                web_site = BeautifulSoup(
                    "".join(map(str, soup.find("div", class_="prop_item site").select('a[href^="https://"]'))),
                    "html.parser")
            except Exception:
                web_site = ""
            try:
                soc_media_vk = BeautifulSoup(
                    "".join(map(str, soup.find("div", {"class": "props_area"}).select('a[href^="https://vk.com/"]'))),
                    "html.parser")
            except Exception:
                soc_media_vk = ""
            try:
                soc_media_you_tube = BeautifulSoup(
                    "".join(
                        map(str, soup.find("div", {"class": "props_area"}).select('a[href^="https://www.youtube.com/"]'))),
                    "html.parser")
            except Exception:
                soc_media_you_tube = ""
            try:
                soup.find(string=re.compile('О компании'))
                about_company = soup.find(string=re.compile('О компании')).find_next()
            except Exception:
                about_company = ""
            try:
                soup.find(string=re.compile('Общая'))
                total_info = soup.find(string=re.compile('Общая')).find_next()
            except Exception:
                total_info = ""

            data = [companies_name, col_number,
                    e_mail, web_site, soc_media_vk, soc_media_you_tube,
                    about_company, total_info]

            # получаем текст из них

            if type(data[0]) == str or None:
                companies_name_text = ""
            else:
                companies_name_text = data[0].text.strip()
            if data[1] == "" or type(data[1]) == str or type(data[1]) is None:
                col_number_text = ""
            else:
                col_number_text = data[1].text.strip()
            if type(data[2]) == str or None:
                e_mail_text = ""
            else:
                e_mail_text = data[2].text.strip()
            if type(data[3]) == str or None:
                web_site_text = ""
            else:
                web_site_text = data[3].text.strip()
            if type(data[4]) == str or None:
                soc_media_vk_text = ""
            else:
                soc_media_vk_text = data[4].text.strip()
            if type(data[5]) == str or None:
                soc_media_you_tube_text = ""
            else:
                soc_media_you_tube_text = data[5].text.strip()
            if data[6] == "" or type(data[6]) == str or type(data[6]) is None:
                about_company_text = ""
            else:
                about_company_text = data[6].text.strip()
            if data[7] == "" or type(data[7]) == str or type(data[7]) is None:
                total_info_text = ""
            else:
                total_info_text = data[7].text.strip()

            wb = openpyxl.load_workbook('РОССИЯ.xlsx')
            ws = wb[f'{item_sheets_name}']

            for i in range(1, 2):
                ws.cell(row=row, column=1, value=companies_name_text)
                ws.cell(row=row, column=2, value=col_number_text)
                ws.cell(row=row, column=3, value=e_mail_text)
                ws.cell(row=row, column=4, value=web_site_text)
                ws.cell(row=row, column=5, value=soc_media_vk_text)
                ws.cell(row=row, column=6, value=soc_media_you_tube_text)
                ws.cell(row=row, column=7, value=about_company_text)
                ws.cell(row=row, column=8, value=total_info_text)
                row += 1

            wb.save('РОССИЯ.xlsx')

            count += 1
            print(f"Итерация № {count}. Компания {company_name} ")
            iteration_count -= 1

            #os.remove(f"data/{company_name}.html")
            #os.remove(f"pages_of_segment/{item_sheets_name}_{number_of_urls_in_this_segment - 1}.html")

            if iteration_count == 0:
                print(f"Работа завершена со страницей: {number_of_urls_in_this_segment}")
                print("---------------------------------")
                if cn_pagination != number_of_urls_in_this_segment:
                    print(f"Приступаем к странице №{number_of_urls_in_this_segment + 1} "
                      f"сегмента '{item_sheets_name}'")
                number_of_urls_in_this_segment += 1
            if iteration_count != 0:
                print(f"Осталось итераций: {iteration_count}")

        #os.remove(f"pages_of_segment/{item_sheets_name}_{number_of_urls_in_this_segment - 1}.html")
        #os.remove(f"data/all_href_companies_of_pages_{item_sheets_name}_{number_of_urls_in_this_segment - 1}.json")
print("Парсер закончил свою работу!")
